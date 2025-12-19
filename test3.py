import pandas as pd
from openpyxl import load_workbook

# Path to the file
path = "ب*******-الفصل الثاني.xlsx"

# Constants
EVALUATION = "معدل تقويم النشاطات /20"
ASSIGNMENT = "الفرض /20"
EXAM = "الإختبار /20"
NOTES = "التقديرات"

# Notes
UNSATISFACTORY = "نتائج غير مرضية اعمل (ي) أكثر"
BELOWAVERAGE = "نتائج دون الوسط يمكنك تحقيق الأفضل"
ACCEPTABLE = "نتائج مقبولة"
GOOD = "نتائج حسنة"
VERYGOOD = "نتائج جيدة"
EXCELLENT = "نتائج جيدة جدا"
VERYEXCELLENT = "عمل ممتاز واصل (ي)"


# Function to calculate score
def rate(x, y, z):
    return (x * 2 + (y + z) / 2) / 3


# Load the workbook (preserves format)
workbook = load_workbook(path)
sheet_names_list = workbook.sheetnames

for sheet in sheet_names_list:
    df = pd.read_excel(path, sheet_name=sheet, skiprows=7)
    df.columns = df.columns.str.strip()  # Clean column names

    # ✅ Check if the NOTES column exists
    if NOTES in df.columns:
        for index, row in df.iterrows():
            # ✅ Directly calculate the score without adding the column
            score = rate(row[EXAM], row[ASSIGNMENT], row[EVALUATION])

            if pd.isna(row[NOTES]):
                if 0 < score < 7:
                    note = UNSATISFACTORY
                elif 7 < score < 10:
                    note = BELOWAVERAGE
                elif 10 < score < 11:
                    note = ACCEPTABLE
                elif 11 < score < 13:
                    note = GOOD
                elif 13 < score < 15:
                    note = VERYGOOD
                elif 15 < score < 17:
                    note = EXCELLENT
                else:
                    note = VERYEXCELLENT

                # ✅ Write the note directly in the existing "NOTES" column
                row_num = index + 2 + 7
                col_num = df.columns.get_loc(NOTES) + 1
                workbook[sheet].cell(row=row_num, column=col_num, value=note)

# ✅ Save changes (preserves format)
workbook.save(path)
