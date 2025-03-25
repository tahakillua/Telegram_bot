import pandas as pd
from openpyxl import load_workbook

path = "بوراس ياسين طه رضوان-الفصل الثاني.xlsx"
# Column
EVALUATION = "معدل تقويم النشاطات /20"
ASSIGNMENT = "الفرض /20"
EXAM = "الإختبار /20"
NOTES = "التقديرات"


#notes

UNSATISFACTORY = "نتائج غير مرضية اعمل (ي) أكثر"
BELOWAVERAGE = "نتائج دون الوسط يمكنك تحقيق الأفضل"
ACCEPTABLE = "نتائج مقبولة"
GOOD = "نتائج حسنة"
VERYGOOD = "نتائج جيدة"
EXCELLENT ="نتائج جيدة جدا"
VERYEXCELLENT ="عمل ممتاز واصل (ي)"
def rate(x, y, z):
    return (x * 2 + (y + z) / 2) / 3

class fillnotes():
    def __init__(self,path):
        self.path = path
        self.file = load_workbook(self.path)

    def notes(self, sheet):
        work_file = pd.read_excel(self.path, sheet_name=sheet, skiprows=7)
        for index, row in work_file.iterrows():
            score = rate(row[EXAM], row[ASSIGNMENT], row[EVALUATION])

            if 0 < score < 7:
                note = UNSATISFACTORY
            elif 7 < score < 10:
                note = BELOWAVERAGE
            elif 10 < score < 11:
                note = ACCEPTABLE
            elif 11 < score < 13:
                note = GOOD
            elif 13 < score < 15:
                note = VERYGOOD
            elif 15 < score < 17:
                note = EXCELLENT
            else:
                note = VERYEXCELLENT
            row_num = index + 2 + 7
            col_num = work_file.columns.get_loc(NOTES) + 1
            self.file[sheet].cell(row=row_num, column=col_num, value=note)

    def savefile(self):
        self.file.save(self.path)

